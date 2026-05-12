"""
Extract per-file fixed-width layouts from each release's Data_Layout XLSX
and emit ONE consolidated CSV per release:

    clean/layouts/{release_id}.csv

Columns: release_id, file_key, srl, field_name, name, block, item, length,
         byte_start, byte_end, remarks

This replaces the older per-file layout CSV scheme (4 files per annual
release, 2 per calendar release) with a single file per release that's
faster to grep, easier to load, and trivial to inspect.

Usage:
    python3 scripts/build_layouts.py                # all releases
    python3 scripts/build_layouts.py annual_2023_24 calendar_2025
"""

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from releases import RELEASES, ROOT
from canonicalize import canonical_name, slug


LAYOUTS_DIR = ROOT / "clean" / "layouts"


# ---- XLSX → in-memory layout rows ----------------------------------------

def _section_rows(ws, start_row: int, end_row: int):
    """Yield rows from a 'Data Layout' section. Skips header/blank rows.

    Accepts srl-less footer rows that have populated length + byte positions
    (some PLFS layouts have a generated weight row without a Srl number).
    """
    fallback_srl = 0
    for r in range(start_row, end_row + 1):
        srl = ws.cell(r, 1).value
        length = ws.cell(r, 5).value
        bs = ws.cell(r, 6).value
        be = ws.cell(r, 7).value
        is_data_row = (
            isinstance(srl, int)
            or (
                isinstance(length, int) and length > 0
                and isinstance(bs, int) and isinstance(be, int)
            )
        )
        if not is_data_row:
            continue
        if not isinstance(srl, int):
            srl = fallback_srl + 1
        fallback_srl = srl
        yield {
            "srl": srl,
            "name": ws.cell(r, 2).value,
            "block": ws.cell(r, 3).value,
            "item": ws.cell(r, 4).value,
            "length": length,
            "byte_start": bs,
            "byte_end": be,
            "remarks": ws.cell(r, 8).value,
        }


def _self_contained_rows(ws, byte_start_col, byte_end_col, fieldname_col):
    """For CY2025-style layouts: per-file sheet carries byte positions AND
    field names directly. No need to cross-reference Data Layout."""
    rows = []
    for r in range(1, ws.max_row + 1):
        srl = ws.cell(r, 1).value
        if not isinstance(srl, int):
            continue
        rows.append({
            "srl": srl,
            "field_name": ws.cell(r, fieldname_col).value,
            "name": ws.cell(r, 2).value,
            "block": ws.cell(r, 3).value,
            "item": ws.cell(r, 4).value,
            "length": ws.cell(r, 5).value,
            "byte_start": ws.cell(r, byte_start_col).value,
            "byte_end": ws.cell(r, byte_end_col).value,
            "remarks": ws.cell(r, 9).value if ws.max_column >= 9 else "",
        })
    return rows


def _attach_field_names_by_srl(rows, fieldname_ws, fieldname_col):
    """For releases where the per-file sheet provides field_name keyed by Srl."""
    names = {}
    for r in range(1, fieldname_ws.max_row + 1):
        s = fieldname_ws.cell(r, 1).value
        if isinstance(s, int):
            names[s] = fieldname_ws.cell(r, fieldname_col).value
    for row in rows:
        row["field_name"] = names.get(row["srl"], "")


def _attach_field_names_by_order(rows, fieldname_ws, fieldname_col, length_col=4):
    """For releases where the per-file sheet has no Srl column — match by row order."""
    ordered = []
    for r in range(1, fieldname_ws.max_row + 1):
        L = fieldname_ws.cell(r, length_col).value
        if isinstance(L, int) and L > 0:
            ordered.append(fieldname_ws.cell(r, fieldname_col).value)
    if len(ordered) != len(rows):
        print(f"  WARN: ordinal-merge size mismatch ({len(rows)} layout, {len(ordered)} names)")
    for row, fn in zip(rows, ordered):
        row["field_name"] = fn


# ---- Disambiguation for per-day CWS fields -------------------------------

def _disambiguate(rows):
    """Block 6 has several Full Names repeated across items 5/3.1 … 5/3.7
    (7 days × 2 activities). Give them mnemonic names like ind11, hrs23, etc.,
    matching CY2024's convention.
    """
    for r in rows:
        name = (r.get("name") or "").lower()
        item = str(r.get("item") or "")
        block = str(r.get("block") or "")
        if block != "6" or "/" not in item:
            continue
        try:
            day = int(item.split("/")[1].split(".")[1])
        except (IndexError, ValueError):
            continue
        activity = 1 if "activity 1" in name else (2 if "activity 2" in name else None)
        if activity and "industry" in name:
            r["field_name"] = f"ind{activity}{day}"
        elif activity and "occupation" in name:
            r["field_name"] = f"ocu{activity}{day}"
        elif activity and "status code" in name:
            r["field_name"] = f"sts{activity}{day}"
        elif activity and "wage earning" in name:
            r["field_name"] = f"ern{activity}{day}"
        elif activity and "hours" in name and "actu" in name:
            r["field_name"] = f"hrs{activity}{day}"
        elif "total hours" in name:
            r["field_name"] = f"tot_hrs{day}"
        elif "hours avail" in name:
            r["field_name"] = f"hav{day}"
        elif "duration of engagement" in name:
            r["field_name"] = f"dur_eng{day}"

    # Block 5.3: duration_of_engagement (principal vs subsidiary share the name)
    for r in rows:
        if str(r.get("block")) == "5.3" and "duration of engagement" in (r.get("name") or "").lower():
            n = r["name"].lower()
            if "principal" in n: r["field_name"] = "dur_pas"
            elif "subsidiary" in n: r["field_name"] = "dur_sas"

    # Block 6: in older layouts both ern_reg and ern_self share "Earnings For Regular..."
    # (PLFS source typo). Use Item to differentiate.
    for r in rows:
        if str(r.get("block")) == "6" and "regular" in (r.get("name") or "").lower():
            it = str(r.get("item") or "")
            if it == "9":  r["field_name"] = "ern_reg"
            elif it == "10": r["field_name"] = "ern_self"


# ---- One release → one consolidated layout CSV ---------------------------

def build_release(release_id: str) -> dict:
    """Build the layout for one release. Returns summary dict."""
    cfg = RELEASES[release_id]
    wb = openpyxl.load_workbook(cfg["xlsx"], data_only=True)

    all_rows: list[dict] = []
    summary = {"release_id": release_id, "files": [], "ok": True}

    for fc in cfg["files"]:
        key = fc["key"]
        if cfg.get("input_kind") in ("csv", "tsv"):
            # Layout XLSX has only the consolidated 'Data Layout' sheet —
            # derive field_name via canonicalize() applied to the Full Name.
            ws = wb["Data Layout"]
            rows = list(_section_rows(ws, *fc["section"]))
            for r in rows:
                cn = canonical_name(r["name"], r.get("block"), r.get("item"))
                r["field_name"] = cn or slug(r["name"])
            _disambiguate(rows)
        elif cfg.get("self_contained"):
            # CY2025 mode — per-file sheet has everything (byte positions + field names)
            ws = wb[fc["fieldname_sheet"]]
            rows = _self_contained_rows(
                ws,
                byte_start_col=fc["byte_start_col"],
                byte_end_col=fc["byte_end_col"],
                fieldname_col=fc["fieldname_col"],
            )
        else:
            # CY2024 mode — Data Layout has byte positions, per-file sheet has field names
            layout_ws = wb["Data Layout"]
            rows = list(_section_rows(layout_ws, *fc["section"]))
            ws = wb[fc["fieldname_sheet"]]
            if fc["srl_join"]:
                _attach_field_names_by_srl(rows, ws, fc["fieldname_col"])
            else:
                _attach_field_names_by_order(rows, ws, fc["fieldname_col"])

        total_bytes = sum(r["length"] or 0 for r in rows)
        ok = total_bytes == fc["byte_total"]
        summary["files"].append({
            "key": key, "n_fields": len(rows),
            "total_bytes": total_bytes,
            "expected_bytes": fc["byte_total"],
            "ok": ok,
        })
        summary["ok"] = summary["ok"] and ok

        for r in rows:
            all_rows.append({
                "release_id": release_id,
                "file_key": key,
                **r,
            })

    # Write the consolidated CSV for this release
    LAYOUTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = LAYOUTS_DIR / f"{release_id}.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "release_id", "file_key", "srl", "field_name", "name",
            "block", "item", "length", "byte_start", "byte_end", "remarks",
        ])
        for r in all_rows:
            fn = (r.get("field_name") or "").strip().lower()
            w.writerow([
                r["release_id"], r["file_key"], r["srl"], fn, r["name"],
                r["block"] or "", r["item"] or "",
                r["length"], r["byte_start"], r["byte_end"],
                r["remarks"] or "",
            ])
    summary["out_path"] = out_path
    return summary


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("releases", nargs="*", choices=list(RELEASES),
                    help="Releases to build; default: all")
    args = ap.parse_args()
    targets = args.releases or list(RELEASES)

    for rid in targets:
        s = build_release(rid)
        print(f"\n=== {rid} → {s['out_path'].relative_to(ROOT)} ===")
        for fr in s["files"]:
            tag = "✓" if fr["ok"] else f"WARN got {fr['total_bytes']} expected {fr['expected_bytes']}"
            print(f"  {fr['key']:<8} {fr['n_fields']:>4} fields  {fr['total_bytes']:>4} bytes  {tag}")


if __name__ == "__main__":
    main()
