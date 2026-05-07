"""
Extract per-file fixed-width layouts from each release's Data_Layout XLSX.

Outputs one CSV per data file into {out_dir}/layout/{file_key}_layout.csv,
each row = one field with srl, name, block, item, length, byte_start,
byte_end, field_name, remarks.

Usage:
    python3 scripts/build_layouts.py                # all known releases
    python3 scripts/build_layouts.py annual_2023_24 # single release
"""

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from releases import RELEASES, ROOT


def extract_data_layout_section(ws, start_row, end_row):
    """Yield rows from 'Data Layout' sheet between section bounds.

    Skips file-header rows and the column-header row (anything where col A
    is not an integer Srl).
    """
    for r in range(start_row, end_row + 1):
        srl = ws.cell(r, 1).value
        if not isinstance(srl, int):
            continue
        yield {
            "srl": srl,
            "name": ws.cell(r, 2).value,
            "block": ws.cell(r, 3).value,
            "item": ws.cell(r, 4).value,
            "length": ws.cell(r, 5).value,
            "byte_start": ws.cell(r, 6).value,
            "byte_end": ws.cell(r, 7).value,
            "remarks": ws.cell(r, 8).value,
        }


def extract_field_names_by_srl(ws, fieldname_col):
    """Map srl -> field_name from a per-file sheet (Srl is in col A)."""
    out = {}
    for r in range(1, ws.max_row + 1):
        srl = ws.cell(r, 1).value
        if not isinstance(srl, int):
            continue
        out[srl] = ws.cell(r, fieldname_col).value
    return out


def extract_field_names_in_order(ws, fieldname_col, length_col=4):
    """Return [field_name, ...] from a per-file sheet that has no Srl column.

    Skips header/title rows (anything where length_col isn't a positive int).
    """
    out = []
    for r in range(1, ws.max_row + 1):
        L = ws.cell(r, length_col).value
        if not (isinstance(L, int) and L > 0):
            continue
        out.append(ws.cell(r, fieldname_col).value)
    return out


def extract_self_contained(ws, byte_start_col, byte_end_col, fieldname_col):
    """Pull the entire layout from a per-file sheet that has byte positions."""
    rows = []
    for r in range(1, ws.max_row + 1):
        srl = ws.cell(r, 1).value
        if not isinstance(srl, int):
            continue
        rows.append(
            {
                "srl": srl,
                "field_name": ws.cell(r, fieldname_col).value,
                "name": ws.cell(r, 2).value,
                "block": ws.cell(r, 3).value,
                "item": ws.cell(r, 4).value,
                "length": ws.cell(r, 5).value,
                "byte_start": ws.cell(r, byte_start_col).value,
                "byte_end": ws.cell(r, byte_end_col).value,
                "remarks": ws.cell(r, 9).value if ws.max_column >= 9 else "",
            }
        )
    return rows


def write_layout(out_path: Path, rows):
    """Write the layout CSV. Field names are normalised to lower-case so column
    names are consistent across releases (the source XLSX is mixed case)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "srl", "field_name", "name", "block", "item",
                "length", "byte_start", "byte_end", "remarks",
            ]
        )
        total = 0
        for r in rows:
            fn = (r.get("field_name") or "").strip().lower()
            w.writerow([
                r["srl"], fn, r["name"],
                r["block"] or "", r["item"] or "",
                r["length"], r["byte_start"], r["byte_end"],
                r["remarks"] or "",
            ])
            total += r["length"] or 0
    return total, len(list(rows))


def write_state_codes(wb, out_path: Path):
    """Pull the State code sheet (if present) into codemaps/state.csv."""
    sheet_names = [s.lower() for s in wb.sheetnames]
    candidates = ("state code", "state codes", "state")
    for cand in candidates:
        if cand in sheet_names:
            ws = wb[wb.sheetnames[sheet_names.index(cand)]]
            with out_path.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["state_code", "state_name"])
                for r in range(1, ws.max_row + 1):
                    code = ws.cell(r, 1).value
                    name = ws.cell(r, 2).value
                    if code is None or name is None:
                        continue
                    code_s = str(code).strip()
                    if not code_s.isdigit():
                        continue
                    w.writerow([code_s.zfill(2), str(name).strip()])
            return True
    return False


def build_release(release_name: str):
    cfg = RELEASES[release_name]
    print(f"\n=== {release_name} — {cfg['label']} ===")
    wb = openpyxl.load_workbook(cfg["xlsx"], data_only=True)

    summary = []
    for fc in cfg["files"]:
        key = fc["key"]
        if cfg["self_contained"]:
            ws = wb[fc["fieldname_sheet"]]
            rows = extract_self_contained(
                ws,
                byte_start_col=fc["byte_start_col"],
                byte_end_col=fc["byte_end_col"],
                fieldname_col=fc["fieldname_col"],
            )
        else:
            layout_ws = wb["Data Layout"]
            section_rows = list(
                extract_data_layout_section(layout_ws, *fc["section"])
            )
            ws = wb[fc["fieldname_sheet"]]
            if fc["srl_join"]:
                names = extract_field_names_by_srl(ws, fc["fieldname_col"])
                for r in section_rows:
                    r["field_name"] = names.get(r["srl"], "")
            else:
                ordered = extract_field_names_in_order(
                    ws, fc["fieldname_col"]
                )
                if len(ordered) != len(section_rows):
                    print(
                        f"  WARN {key}: ordinal merge size mismatch "
                        f"(layout={len(section_rows)} fields={len(ordered)})"
                    )
                for r, fn in zip(section_rows, ordered):
                    r["field_name"] = fn
            rows = section_rows

        out_path = cfg["layout_dir"] / f"{key}_layout.csv"
        total_bytes, n = write_layout(out_path, rows)
        ok = "✓" if total_bytes == fc["byte_total"] else f"WARN got {total_bytes} expected {fc['byte_total']}"
        summary.append((key, n, total_bytes, fc["byte_total"], ok))
        print(f"  {key:<8} {n:>4} fields  {total_bytes:>4} bytes  expected {fc['byte_total']:>4}  {ok}")

    # State codes (only annual XLSX has a 'State code' sheet — calendar layouts
    # may or may not).
    state_csv = ROOT / "codemaps" / "state.csv"
    if not state_csv.exists():
        if write_state_codes(wb, state_csv):
            print(f"  state codes -> {state_csv.relative_to(ROOT)}")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "releases",
        nargs="*",
        choices=list(RELEASES),
        help="Release(s) to build; default: all known releases",
    )
    args = ap.parse_args()
    targets = args.releases or list(RELEASES)
    for r in targets:
        build_release(r)


if __name__ == "__main__":
    main()
